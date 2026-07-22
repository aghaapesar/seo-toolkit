"""
Excel export / re-import for technical SEO audit reports.

Input:
    Audit result dict (issues with full `urls` lists) + report_id.

Output:
    Styled .xlsx workbooks (technical / content / all) with status column,
    conditional green rows when marked done, and parsers for re-check uploads.
"""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

logger = logging.getLogger(__name__)

# Status values shown in the dropdown (Persian UI)
STATUS_OPEN = "☐ باز"
STATUS_DONE = "☑ انجام‌شده"
STATUS_CHOICES = f'"{STATUS_OPEN},{STATUS_DONE}"'

# Owner → workbook bucket
OWNER_TECH = "تیم فنی"
OWNER_CONTENT = "تیم محتوا"

# Column headers (Persian) — order matters for parser
HEADERS = [
    "وضعیت",
    "اولویت",
    "شدت",
    "دسته",
    "عنوان مشکل",
    "آدرس صفحه",
    "تعداد صفحات",
    "توضیح",
    "پیشنهاد اصلاح",
    "راهکار استک",
    "مسئول",
    "تلاش",
    "یادداشت",
    "شناسه مشکل",
    "شناسه گزارش",
]

COL_STATUS = 1
COL_PRIORITY = 2
COL_SEVERITY = 3
COL_CATEGORY = 4
COL_TITLE = 5
COL_URL = 6
COL_COUNT = 7
COL_DESC = 8
COL_FIX = 9
COL_STACK = 10
COL_OWNER = 11
COL_EFFORT = 12
COL_NOTES = 13
COL_ISSUE_ID = 14
COL_REPORT_ID = 15

# Column widths (readable task board)
COLUMN_WIDTHS = {
    COL_STATUS: 14,
    COL_PRIORITY: 8,
    COL_SEVERITY: 10,
    COL_CATEGORY: 14,
    COL_TITLE: 36,
    COL_URL: 48,
    COL_COUNT: 12,
    COL_DESC: 40,
    COL_FIX: 40,
    COL_STACK: 36,
    COL_OWNER: 12,
    COL_EFFORT: 8,
    COL_NOTES: 22,
    COL_ISSUE_ID: 22,
    COL_REPORT_ID: 28,
}

SEVERITY_FILL = {
    "critical": "FECACA",
    "high": "FED7AA",
    "medium": "FEF08A",
    "low": "BBF7D0",
}

SEVERITY_ORDER = ("critical", "high", "medium", "low")


def _thin_border() -> Border:
    side = Side(style="thin", color="CBD5E1")
    return Border(left=side, right=side, top=side, bottom=side)


def _is_done_status(value: Any) -> bool:
    """True when Excel status cell means the task is completed."""
    text = str(value or "").strip().lower()
    if not text:
        return False
    markers = (
        "انجام",
        "done",
        "fixed",
        "completed",
        "☑",
        "✓",
        "✔",
        "true",
        "yes",
        "1",
    )
    return any(m in text for m in markers) and "باز" not in text


def _owner_bucket(owner: str) -> str:
    """
    Map issue owner to export bucket key.

    Output:
        'technical' | 'content' | 'other'
    """
    owner = (owner or "").strip()
    if OWNER_CONTENT in owner or "محتوا" in owner:
        return "content"
    if OWNER_TECH in owner or "فنی" in owner:
        return "technical"
    return "other"


def flatten_issue_rows(
    audit_result: Dict[str, Any],
    report_id: str,
) -> List[Dict[str, Any]]:
    """
    Expand aggregated issues into one Excel row per affected URL.

    Input:
        audit_result: Auditor result with `issues` list.
        report_id: Stable id written into each row for re-upload matching.

    Output:
        Row dicts ready for workbook writing (already priority-sorted).
    """
    issues = list(audit_result.get("issues") or [])
    sev_rank = {s: i for i, s in enumerate(SEVERITY_ORDER)}
    issues.sort(
        key=lambda i: (
            sev_rank.get(str(i.get("severity") or ""), 9),
            -int(i.get("count") or 0),
        )
    )

    rows: List[Dict[str, Any]] = []
    priority = 0
    for issue in issues:
        urls = list(issue.get("urls") or issue.get("sample_urls") or [])
        if not urls:
            # Site-level issues still need a trackable row
            urls = [str(audit_result.get("site_url") or "")]
        for url in urls:
            priority += 1
            rows.append(
                {
                    "status": STATUS_OPEN,
                    "priority": priority,
                    "severity": issue.get("severity") or "",
                    "severity_fa": issue.get("severity_fa")
                    or issue.get("severity")
                    or "",
                    "category": issue.get("category") or "",
                    "title": issue.get("title") or "",
                    "url": url or "",
                    "count": int(issue.get("count") or len(urls)),
                    "description": issue.get("description") or "",
                    "recommendation": issue.get("recommendation") or "",
                    "stack_solution": issue.get("stack_solution") or "",
                    "owner": issue.get("owner") or "",
                    "effort": issue.get("effort") or "",
                    "notes": "",
                    "issue_id": issue.get("issue_id") or "",
                    "report_id": report_id,
                }
            )
    return rows


def _write_guide_sheet(wb: Workbook, team_label: str) -> None:
    """Add a Persian how-to sheet for the task workbook."""
    ws = wb.create_sheet("راهنما", 0)
    lines = [
        f"کاربرگ پیگیری مشکلات — {team_label}",
        "",
        "نحوه استفاده:",
        "۱. ستون «وضعیت» را از لیست انتخاب کنید: ☐ باز یا ☑ انجام‌شده.",
        "۲. وقتی وضعیت را روی انجام‌شده بگذارید، کل ردیف سبز می‌شود.",
        "۳. با فیلتر بالای جدول می‌توانید فقط موارد باز یا یک شدت خاص را ببینید.",
        "۴. ستون «یادداشت» برای توضیح تیم آزاد است.",
        "۵. پس از اصلاح، همین فایل را در صفحه «بررسی مشکلات فنی» آپلود کنید تا سیستم موارد باز را دوباره چک کند.",
        "",
        "نکته: ستون‌های «شناسه مشکل» و «شناسه گزارش» را پاک نکنید — برای تطبیق خودکار لازم‌اند.",
    ]
    ws.column_dimensions["A"].width = 100
    title_font = Font(name="Arial", size=14, bold=True, color="0F172A")
    body_font = Font(name="Arial", size=11, color="334155")
    for idx, line in enumerate(lines, start=1):
        cell = ws.cell(row=idx, column=1, value=line)
        cell.font = title_font if idx == 1 else body_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 24


def _style_tasks_sheet(ws, row_count: int) -> None:
    """
    Apply header style, borders, freeze, filter, status dropdown, green-on-done CF.

    Input:
        ws: openpyxl worksheet with header + data already written.
        row_count: Number of data rows (excluding header).
    """
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = _thin_border()

    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS.get(col, 16)

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADERS))
    last_row = max(1, row_count + 1)
    ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    # Status dropdown
    if row_count > 0:
        dv = DataValidation(
            type="list",
            formula1=STATUS_CHOICES,
            allow_blank=False,
            showDropDown=False,
            showErrorMessage=True,
            errorTitle="وضعیت نامعتبر",
            error="فقط ☐ باز یا ☑ انجام‌شده را انتخاب کنید.",
        )
        dv.add(f"A2:A{last_row}")
        ws.add_data_validation(dv)

        # Whole-row green when status cell contains «انجام»
        green = PatternFill("solid", fgColor="BBF7D0")
        green_font = Font(name="Arial", size=10, color="14532D")
        # Relative formula: $A2 checked for each row in the range
        formula = f'ISNUMBER(SEARCH("انجام",$A2))'
        ws.conditional_formatting.add(
            f"A2:{last_col}{last_row}",
            FormulaRule(formula=[formula], fill=green, font=green_font),
        )

    # Per-row base styling + severity tint on severity column
    data_font = Font(name="Arial", size=10, color="0F172A")
    data_align = Alignment(vertical="top", wrap_text=True, readingOrder=2)
    zebra = PatternFill("solid", fgColor="F8FAFC")

    for r in range(2, last_row + 1):
        for c in range(1, len(HEADERS) + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = border
            if r % 2 == 0 and c != COL_SEVERITY:
                cell.fill = zebra
        sev_cell = ws.cell(row=r, column=COL_SEVERITY)
        # Match severity from hidden issue context via priority row's Persian label —
        # also try English key if present in adjacent note; prefer fill from title context.
        # Severity column holds Persian label; map via known FA labels.
        sev_text = str(sev_cell.value or "")
        sev_key = ""
        for key, label in (
            ("critical", "بحرانی"),
            ("high", "بالا"),
            ("medium", "متوسط"),
            ("low", "پایین"),
        ):
            if label in sev_text or key == sev_text.lower():
                sev_key = key
                break
        if sev_key in SEVERITY_FILL:
            sev_cell.fill = PatternFill("solid", fgColor=SEVERITY_FILL[sev_key])

    # RTL for Persian readability
    ws.sheet_view.rightToLeft = True


def write_team_workbook(
    rows: Sequence[Dict[str, Any]],
    output_path: Path,
    *,
    team_label: str,
) -> Path:
    """
    Write one styled Excel workbook for a team bucket.

    Input:
        rows: Flattened issue/URL rows (already filtered for this team).
        output_path: Destination .xlsx path.
        team_label: Persian label for guide sheet.

    Output:
        Path to the written file.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    # Remove default sheet after guide + tasks are ready
    default = wb.active
    wb.remove(default)
    _write_guide_sheet(wb, team_label)

    ws = wb.create_sheet("پیگیری مشکلات", 1)
    for col, header in enumerate(HEADERS, start=1):
        ws.cell(row=1, column=col, value=header)

    for r_idx, row in enumerate(rows, start=2):
        values = [
            row.get("status", STATUS_OPEN),
            row.get("priority"),
            row.get("severity_fa") or row.get("severity"),
            row.get("category"),
            row.get("title"),
            row.get("url"),
            row.get("count"),
            row.get("description"),
            row.get("recommendation"),
            row.get("stack_solution"),
            row.get("owner"),
            row.get("effort"),
            row.get("notes"),
            row.get("issue_id"),
            row.get("report_id"),
        ]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)

    _style_tasks_sheet(ws, len(rows))
    wb.save(output_path)
    logger.info("Wrote audit Excel: %s (%s rows)", output_path, len(rows))
    return output_path


def generate_audit_excels(
    audit_result: Dict[str, Any],
    out_dir: Path,
    report_id: str,
) -> Dict[str, Path]:
    """
    Build technical / content / all Excel workbooks next to the PDF/JSON.

    Input:
        audit_result: Full audit JSON-serializable result.
        out_dir: technical_audit output directory.
        report_id: Base id without extension (e.g. audit_zitro-ir_20260721_120000).

    Output:
        Dict of label → Path for each generated workbook.
    """
    all_rows = flatten_issue_rows(audit_result, report_id)

    def _copy_bucket(key: str) -> List[Dict[str, Any]]:
        """Shallow-copy rows for one owner bucket and re-number priority."""
        copied = [
            dict(row)
            for row in all_rows
            if _owner_bucket(str(row.get("owner") or "")) == key
        ]
        for i, row in enumerate(copied, start=1):
            row["priority"] = i
        return copied

    files: Dict[str, Path] = {}
    mapping = (
        ("technical", "فنی", OWNER_TECH, _copy_bucket("technical")),
        ("content", "محتوا", OWNER_CONTENT, _copy_bucket("content")),
        ("all", "همه", "همه تیم‌ها", [dict(r) for r in all_rows]),
    )

    for key, fa_suffix, label, rows in mapping:
        if key != "all" and not rows:
            continue
        path = out_dir / f"{report_id}_{fa_suffix}.xlsx"
        write_team_workbook(rows, path, team_label=label)
        files[key] = path

    return files


def parse_status_workbook(path: Path) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Parse an uploaded tracking Excel.

    Input:
        path: .xlsx file path.

    Output:
        (report_id, rows) where each row has issue_id, url, status, is_done, …
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    # Prefer the tracking sheet; else first sheet with our headers
    ws = None
    for name in wb.sheetnames:
        if "پیگیری" in name or name.lower() in ("tasks", "issues"):
            ws = wb[name]
            break
    if ws is None:
        ws = wb[wb.sheetnames[-1]] if wb.sheetnames else None
    if ws is None:
        raise ValueError("Excel workbook has no sheets")

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise ValueError("Excel is empty")

    # Map header names → index (tolerant of slight renames)
    index: Dict[str, int] = {}
    for i, name in enumerate(header):
        key = str(name or "").strip()
        if key:
            index[key] = i

    def col(*aliases: str, default: int) -> int:
        for a in aliases:
            if a in index:
                return index[a]
        return default

    i_status = col("وضعیت", "status", default=0)
    i_url = col("آدرس صفحه", "url", "URL", default=5)
    i_issue = col("شناسه مشکل", "issue_id", default=13)
    i_report = col("شناسه گزارش", "report_id", default=14)
    i_title = col("عنوان مشکل", "title", default=4)
    i_owner = col("مسئول", "owner", default=10)
    i_sev = col("شدت", "severity", default=2)

    parsed: List[Dict[str, Any]] = []
    report_id = ""
    for raw in rows_iter:
        if not raw or all(v is None or str(v).strip() == "" for v in raw):
            continue

        def cell(idx: int) -> str:
            if idx < 0 or idx >= len(raw):
                return ""
            return str(raw[idx] if raw[idx] is not None else "").strip()

        status = cell(i_status)
        issue_id = cell(i_issue)
        url = cell(i_url)
        rid = cell(i_report)
        if rid and not report_id:
            report_id = rid
        if not issue_id and not url:
            continue
        parsed.append(
            {
                "status": status,
                "is_done": _is_done_status(status),
                "issue_id": issue_id,
                "url": url,
                "title": cell(i_title),
                "owner": cell(i_owner),
                "severity_fa": cell(i_sev),
                "report_id": rid,
            }
        )

    wb.close()
    if not report_id and parsed:
        report_id = str(parsed[0].get("report_id") or "")
    return report_id, parsed


def open_issue_pairs(rows: Iterable[Dict[str, Any]]) -> List[Tuple[str, str]]:
    """
    Collect (issue_id, url) pairs that are still open.

    Output:
        Deduplicated list for re-check targeting.
    """
    seen = set()
    out: List[Tuple[str, str]] = []
    for row in rows:
        if row.get("is_done"):
            continue
        issue_id = str(row.get("issue_id") or "").strip()
        url = str(row.get("url") or "").strip()
        if not issue_id:
            continue
        key = (issue_id, url)
        if key in seen:
            continue
        seen.add(key)
        out.append(key)
    return out
